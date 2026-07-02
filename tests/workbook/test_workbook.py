"""Workbook module tests: round-trips, safety rules, header validation, locks, backups."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from testdata.seed_workbook import create_seed_workbook
from triage_agent.workbook.models import Draft, LogEntry, ManualQueueEntry
from triage_agent.workbook.reader import get_all_candidates, get_candidates_by_status, get_config, get_drafts_by_approval, load_workbook
from triage_agent.workbook.schema import (
    CANDIDATES_ALL_COLUMNS,
    TAB_CANDIDATES,
    AdminColumnWriteError,
    CandidateNotFoundError,
    FileLockError,
    HeaderMismatchError,
)
from triage_agent.workbook.writer import append_draft, append_log, append_manual_queue, save_workbook, update_candidate


@pytest.fixture()
def wb_path(tmp_path: Path) -> Path:
    path = tmp_path / "master.xlsx"
    create_seed_workbook(path)
    return path


@pytest.fixture()
def wb_ctx(wb_path: Path):
    return load_workbook(wb_path)


# ---------------------------------------------------------------------------
# Seed
# ---------------------------------------------------------------------------

def test_seed_creates_file(wb_path: Path) -> None:
    assert wb_path.exists()


def test_seed_has_fifty_candidates(wb_ctx) -> None:
    candidates = get_all_candidates(wb_ctx)
    assert len(candidates) == 50


def test_all_seeded_candidates_default_to_new(wb_ctx) -> None:
    new_candidates = get_candidates_by_status(wb_ctx, "New")
    assert len(new_candidates) == 50


# ---------------------------------------------------------------------------
# Round-trip read / write
# ---------------------------------------------------------------------------

def test_round_trip_update_status(wb_path: Path, wb_ctx) -> None:
    update_candidate(wb_ctx, "C001", {"status": "Flagged", "trigger_reason": "post_visit"})
    save_workbook(wb_ctx, wb_path)

    wb2 = load_workbook(wb_path)
    flagged = get_candidates_by_status(wb2, "Flagged")
    still_new = get_candidates_by_status(wb2, "New")

    assert len(flagged) == 1
    assert flagged[0].candidate_id == "C001"
    assert flagged[0].trigger_reason == "post_visit"
    assert len(still_new) == 49


def test_round_trip_flags_list(wb_path: Path, wb_ctx) -> None:
    update_candidate(wb_ctx, "C023", {"flags": ["critical_lab", "elderly"], "urgency_tier": "high"})
    save_workbook(wb_ctx, wb_path)

    wb2 = load_workbook(wb_path)
    all_candidates = get_all_candidates(wb2)
    c023 = next(c for c in all_candidates if c.candidate_id == "C023")
    assert set(c023.flags) == {"critical_lab", "elderly"}
    assert c023.urgency_tier == "high"


def test_round_trip_append_draft(wb_path: Path, wb_ctx) -> None:
    draft = Draft(
        draft_id="D001",
        candidate_id="C001",
        channel="email",
        draft_text="Hi Maria, following up on your recent visit...",
    )
    append_draft(wb_ctx, draft)
    save_workbook(wb_ctx, wb_path)

    wb2 = load_workbook(wb_path)
    pending = get_drafts_by_approval(wb2, "Pending")
    assert len(pending) == 1
    assert pending[0].draft_id == "D001"
    assert pending[0].candidate_id == "C001"


def test_round_trip_append_log(wb_path: Path, wb_ctx) -> None:
    entry = LogEntry(
        timestamp=datetime(2026, 7, 2, 9, 0, 0),
        run_id="run-001",
        action="triage",
        candidate_id="C001",
        detail="Urgency scored: high",
    )
    append_log(wb_ctx, entry)
    save_workbook(wb_ctx, wb_path)

    raw_wb = openpyxl.load_workbook(wb_path)
    ws = raw_wb["Log"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][2] == "triage"
    assert rows[0][3] == "C001"


def test_round_trip_append_manual_queue(wb_path: Path, wb_ctx) -> None:
    entry = ManualQueueEntry(
        candidate_id="C024",
        patient_name="Norman Thompson",
        urgency="high",
        summary="Critically high fasting glucose, patient not reachable.",
        recommended_action="Escalate to on-call nurse",
        flags="critical_lab",
    )
    append_manual_queue(wb_ctx, entry)
    save_workbook(wb_ctx, wb_path)

    raw_wb = openpyxl.load_workbook(wb_path)
    ws = raw_wb["Manual Queue"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][0] == "C024"
    assert rows[0][7] == "No"  # resolved default


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

def test_get_config_parses_flag_vocabulary(wb_ctx) -> None:
    config = get_config(wb_ctx)
    assert config.autosend_enabled is False
    assert "endo" in config.specialty_scope
    assert "abnormal_lab" in config.flag_vocabulary
    assert "critical_lab" in config.flag_vocabulary
    assert len(config.flag_vocabulary) == 10


def test_get_config_temperature_is_float(wb_ctx) -> None:
    config = get_config(wb_ctx)
    assert isinstance(config.temperature, float)
    assert config.temperature == 0.0


# ---------------------------------------------------------------------------
# Safety: refuse admin column writes
# ---------------------------------------------------------------------------

def test_refuses_write_to_admin_column(wb_ctx) -> None:
    with pytest.raises(AdminColumnWriteError) as exc_info:
        update_candidate(wb_ctx, "C001", {"patient_name": "Hacked Name"})
    assert "patient_name" in str(exc_info.value)


def test_refuses_write_to_multiple_admin_columns(wb_ctx) -> None:
    with pytest.raises(AdminColumnWriteError) as exc_info:
        update_candidate(wb_ctx, "C001", {"patient_id": "X", "appointment_date": "2026-01-01"})
    error_msg = str(exc_info.value)
    assert "patient_id" in error_msg or "appointment_date" in error_msg


def test_refuses_mixed_admin_and_agent_columns(wb_ctx) -> None:
    with pytest.raises(AdminColumnWriteError):
        update_candidate(wb_ctx, "C001", {"status": "Flagged", "patient_name": "Hacked"})


def test_raises_on_missing_candidate(wb_ctx) -> None:
    with pytest.raises(CandidateNotFoundError):
        update_candidate(wb_ctx, "DOES_NOT_EXIST", {"status": "Flagged"})


# ---------------------------------------------------------------------------
# Safety: header mismatch detection
# ---------------------------------------------------------------------------

def test_header_mismatch_on_read(tmp_path: Path) -> None:
    path = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = TAB_CANDIDATES
    ws.append(["id", "name", "wrong_column"])
    wb.save(path)

    wb_ctx = load_workbook(path)
    with pytest.raises(HeaderMismatchError):
        get_all_candidates(wb_ctx)


def test_header_mismatch_on_write(tmp_path: Path) -> None:
    path = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = TAB_CANDIDATES
    ws.append(["id", "name"])
    wb.save(path)

    wb_ctx = load_workbook(path)
    with pytest.raises(HeaderMismatchError):
        update_candidate(wb_ctx, "C001", {"status": "Flagged"})


def test_header_mismatch_message_names_tab(tmp_path: Path) -> None:
    path = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = TAB_CANDIDATES
    ws.append(["wrong"])
    wb.save(path)

    wb_ctx = load_workbook(path)
    with pytest.raises(HeaderMismatchError, match="Candidates"):
        get_all_candidates(wb_ctx)


def test_correct_column_order_required(tmp_path: Path) -> None:
    path = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = TAB_CANDIDATES
    ws.append(list(reversed(CANDIDATES_ALL_COLUMNS)))  # same columns, wrong order
    wb.save(path)

    wb_ctx = load_workbook(path)
    with pytest.raises(HeaderMismatchError):
        get_all_candidates(wb_ctx)


# ---------------------------------------------------------------------------
# Safety: file lock detection
# ---------------------------------------------------------------------------

def test_file_lock_detected_via_marker(wb_path: Path, wb_ctx) -> None:
    lock_marker = wb_path.parent / f"~${wb_path.name}"
    lock_marker.touch()
    try:
        with pytest.raises(FileLockError, match="open in another application"):
            save_workbook(wb_ctx, wb_path)
    finally:
        lock_marker.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# Safety: backup creation
# ---------------------------------------------------------------------------

def test_backup_created_on_save(wb_path: Path, wb_ctx) -> None:
    save_workbook(wb_ctx, wb_path)

    backups_dir = wb_path.parent / "backups"
    assert backups_dir.exists()
    backups = list(backups_dir.glob("*.xlsx"))
    assert len(backups) == 1
    assert backups[0].name.startswith("master_")


def test_backup_name_contains_timestamp(wb_path: Path, wb_ctx) -> None:
    save_workbook(wb_ctx, wb_path)

    backups_dir = wb_path.parent / "backups"
    backup = next(backups_dir.glob("*.xlsx"))
    # Name format: master_YYYYMMDD_HHMMSS.xlsx
    parts = backup.stem.split("_")
    assert len(parts) == 3
    assert parts[1].isdigit() and len(parts[1]) == 8
    assert parts[2].isdigit() and len(parts[2]) == 6


def test_no_backup_created_for_new_file(tmp_path: Path) -> None:
    path = tmp_path / "new.xlsx"
    create_seed_workbook(path)
    wb_ctx = load_workbook(path)
    path.unlink()  # remove the file so save_workbook sees it as new

    save_workbook(wb_ctx, path)
    backups_dir = tmp_path / "backups"
    assert not backups_dir.exists()
