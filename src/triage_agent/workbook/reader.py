import logging
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from triage_agent.workbook.models import Candidate, Config, Draft
from triage_agent.workbook.schema import (
    CANDIDATES_ALL_COLUMNS,
    CONFIG_COLUMNS,
    DRAFTS_COLUMNS,
    TAB_CANDIDATES,
    TAB_CONFIG,
    TAB_DRAFTS,
    HeaderMismatchError,
)

logger = logging.getLogger(__name__)


@dataclass
class WorkbookContext:
    path: Path
    wb: openpyxl.Workbook


def load_workbook(path: str | Path) -> WorkbookContext:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    wb = openpyxl.load_workbook(path)
    return WorkbookContext(path=path, wb=wb)


def get_sheet_headers(ws: Worksheet) -> list[str]:
    """Return non-None header values from row 1, stripping trailing empty cells."""
    row = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    while row and row[-1] is None:
        row.pop()
    return row  # type: ignore[return-value]


def validate_sheet_headers(ws: Worksheet, expected: list[str]) -> None:
    """Raise HeaderMismatchError if row 1 does not exactly match expected."""
    actual = get_sheet_headers(ws)
    if actual != expected:
        raise HeaderMismatchError(
            f"Tab '{ws.title}': expected columns {expected}, got {actual}"
        )


def get_all_candidates(wb_ctx: WorkbookContext) -> list[Candidate]:
    ws = wb_ctx.wb[TAB_CANDIDATES]
    validate_sheet_headers(ws, CANDIDATES_ALL_COLUMNS)

    headers = get_sheet_headers(ws)
    results: list[Candidate] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        if not row_dict.get("candidate_id"):
            continue
        results.append(Candidate(**_coerce_candidate(row_dict)))
    return results


def get_candidates_by_status(wb_ctx: WorkbookContext, status: str) -> list[Candidate]:
    return [c for c in get_all_candidates(wb_ctx) if c.status == status]


def get_drafts_by_approval(wb_ctx: WorkbookContext, approval_status: str) -> list[Draft]:
    ws = wb_ctx.wb[TAB_DRAFTS]
    validate_sheet_headers(ws, DRAFTS_COLUMNS)

    headers = get_sheet_headers(ws)
    results: list[Draft] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        if not row_dict.get("draft_id"):
            continue
        if str(row_dict.get("approval_status", "")) == approval_status:
            results.append(Draft(**{k: v for k, v in row_dict.items() if v is not None}))
    return results


def get_config(wb_ctx: WorkbookContext) -> Config:
    ws = wb_ctx.wb[TAB_CONFIG]
    validate_sheet_headers(ws, CONFIG_COLUMNS)

    raw: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key, value = row[0], row[1]
        if key is not None:
            raw[str(key)] = str(value) if value is not None else ""

    return _parse_config(raw)


def _coerce_candidate(row: dict) -> dict:
    flags = row.get("flags")
    if not flags:
        row["flags"] = []
    elif isinstance(flags, str):
        row["flags"] = [f.strip() for f in flags.split(",") if f.strip()]

    if not row.get("status"):
        row["status"] = "New"

    # Strip None values so Pydantic field defaults apply for optional fields
    return {k: v for k, v in row.items() if v is not None or k in ("flags", "status")}


def _parse_config(raw: dict[str, str]) -> Config:
    flag_vocab: dict[str, str] = {}
    other: dict[str, str] = {}

    for key, value in raw.items():
        if key.startswith("flag:"):
            flag_vocab[key[5:]] = value
        else:
            other[key] = value

    return Config(
        autosend_enabled=other.get("autosend_enabled", "False").lower() == "true",
        specialty_scope=[s.strip() for s in other.get("specialty_scope", "").split(",") if s.strip()],
        classifier_model=other.get("classifier_model", "claude-sonnet-4-6"),
        triage_model=other.get("triage_model", "claude-sonnet-4-6"),
        temperature=float(other.get("temperature", "0.0")),
        flag_vocabulary=flag_vocab,
        raw=other,
    )
