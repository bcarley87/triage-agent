"""Generate a seed master.xlsx with realistic Endocrinology data for development and testing."""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from triage_agent.workbook.schema import (
    CANDIDATES_ALL_COLUMNS,
    CONFIG_COLUMNS,
    DRAFTS_COLUMNS,
    LOG_COLUMNS,
    MANUAL_QUEUE_COLUMNS,
    TAB_CANDIDATES,
    TAB_CONFIG,
    TAB_DRAFTS,
    TAB_LOG,
    TAB_MANUAL_QUEUE,
)

FLAG_VOCABULARY: dict[str, str] = {
    "abnormal_lab": "lab value outside the normal range for this test type",
    "critical_lab": "lab value in the emergency-notify range",
    "chronic_patient": "patient has one or more chronic conditions",
    "elderly": "patient is 70 or older",
    "prior_no_show": "patient has missed at least one appointment in the past year",
    "medication_change": "appointment involved starting, stopping, or adjusting a medication",
    "missed_appointment": "patient did not attend their scheduled visit",
    "overdue_monitoring": "scheduled monitoring interval has lapsed",
    "post_procedure": "appointment was a procedure requiring followup",
    "new_diagnosis": "patient received a new diagnosis at this visit",
}

# fmt: off
# (candidate_id, patient_id, patient_name, appointment_date, visit_type, lab_type, lab_value, admin_notes)
_CANDIDATE_ROWS: list[tuple] = [
    # POST-VISIT CHECKINS (12)
    ("C001", "P001", "Maria Santos",      date(2026, 6, 18), "Diabetes Follow-up",         None,              None,        "3-month visit. Metformin 1000mg BID ongoing."),
    ("C002", "P002", "James Liu",         date(2026, 6, 20), "Thyroid Follow-up",           None,              None,        "Levothyroxine dose was adjusted this visit."),
    ("C003", "P003", "Priya Patel",       date(2026, 6, 22), "New Patient Consult",         None,              None,        "New patient, T2DM. Started on Ozempic 0.5mg weekly."),
    ("C004", "P004", "Robert Kim",        date(2026, 6, 10), "Graves Disease Follow-up",    None,              None,        "Post-RAI therapy. Monitoring for hypothyroidism."),
    ("C005", "P005", "Dorothy Williams",  date(2026, 6, 5),  "Adrenal Follow-up",           None,              None,        "Hydrocortisone management, stable. Age 74."),
    ("C006", "P006", "Carlos Rivera",     date(2026, 6, 25), "Diabetes Management",         None,              None,        "Insulin pump review. Carb ratios adjusted."),
    ("C007", "P007", "Sarah Chen",        date(2026, 6, 28), "Hashimoto's Follow-up",       None,              None,        "Dose maintained. Stable thyroid function."),
    ("C008", "P008", "Michael Johnson",   date(2026, 6, 12), "Acromegaly Follow-up",        None,              None,        "Octreotide ongoing. IGF-1 trending down."),
    ("C009", "P009", "Fatima Al-Hassan",  date(2026, 6, 30), "PCOS Consult",                None,              None,        "Metformin started for insulin resistance."),
    ("C010", "P010", "William Thompson",  date(2026, 6, 8),  "Cushing's Follow-up",         None,              None,        "Post-surgical monitoring. Age 71."),
    ("C011", "P011", "Linda Martinez",    date(2026, 6, 15), "Osteoporosis + DM2",          None,              None,        "Calcium and Vit D supplementation reviewed."),
    ("C012", "P012", "Daniel Park",       date(2026, 6, 17), "Hyperthyroidism Follow-up",   None,              None,        "PTU titration visit. Repeat TFTs in 6 weeks."),
    # ABNORMAL LABS (10)
    ("C013", "P013", "Jennifer White",    date(2026, 6, 24), "Lab Review",                  "HbA1c",           "10.2%",     "HbA1c significantly above target of 7%."),
    ("C014", "P014", "Thomas Brown",      date(2026, 6, 23), "Lab Review",                  "TSH",             "0.04 mIU/L","TSH suppressed. On Levothyroxine 150mcg."),
    ("C015", "P015", "Amanda Garcia",     date(2026, 6, 26), "Lab Review",                  "Fasting Glucose", "285 mg/dL", "Fasting glucose very high. Recent illness noted."),
    ("C016", "P016", "Steven Lee",        date(2026, 6, 19), "Lab Review",                  "Free T4",         "0.4 ng/dL", "Low Free T4. Possible non-compliance."),
    ("C017", "P017", "Rachel Moore",      date(2026, 6, 21), "Lab Review",                  "HbA1c",           "9.8%",      "HbA1c worsening from 8.9% six months ago."),
    ("C018", "P018", "Kevin Wilson",      date(2026, 6, 16), "Lab Review",                  "TSH",             "18.2 mIU/L","TSH markedly elevated. Patient reports fatigue."),
    ("C019", "P019", "Jessica Taylor",    date(2026, 6, 27), "Lab Review",                  "AM Cortisol",     "1.2 mcg/dL","Low morning cortisol. Adrenal insufficiency concern."),
    ("C020", "P020", "Brian Anderson",    date(2026, 6, 14), "Lab Review",                  "LDL",             "198 mg/dL", "LDL high in diabetic patient. No statin currently."),
    ("C021", "P021", "Ashley Jackson",    date(2026, 6, 29), "Lab Review",                  "HbA1c",           "11.1%",     "Third consecutive HbA1c above 10%. Age 68."),
    ("C022", "P022", "Gregory Martin",    date(2026, 6, 13), "Lab Review",                  "TSH",             "0.01 mIU/L","TSH critically suppressed on Synthroid + semaglutide."),
    # CRITICAL LABS (3)
    ("C023", "P023", "Eleanor Harris",    date(2026, 7, 1),  "Critical Lab Review",         "HbA1c",           "14.3%",     "Emergency-level HbA1c. Age 72. Possible DKA risk."),
    ("C024", "P024", "Norman Thompson",   date(2026, 7, 1),  "Critical Lab Review",         "Fasting Glucose", "489 mg/dL", "Critically high fasting glucose. Patient not answering calls."),
    ("C025", "P025", "Patricia Robinson", date(2026, 6, 30), "Critical Lab Review",         "AM Cortisol",     "0.4 mcg/dL","Critically low cortisol. Adrenal crisis risk. Age 78."),
    # MISSED APPOINTMENTS (8)
    ("C026", "P026", "David Clark",       date(2026, 6, 18), "Missed: Diabetes 3-Month",    None,              None,        "No show. Second missed appointment in 6 months."),
    ("C027", "P027", "Sandra Lewis",      date(2026, 6, 20), "Missed: Thyroid Follow-up",   None,              None,        "No show. On Levothyroxine, TSH check due."),
    ("C028", "P028", "Christopher Walker",date(2026, 6, 25), "Missed: Annual Diabetes Review",None,            None,        "No show. HbA1c overdue by 3 months."),
    ("C029", "P029", "Michelle Hall",     date(2026, 6, 22), "Missed: HbA1c Review",        None,              None,        "No show. Prior HbA1c was 8.7%."),
    ("C030", "P030", "Joseph Young",      date(2026, 6, 17), "Missed: Medication Review",   None,              None,        "No show. Insulin titration was planned this visit."),
    ("C031", "P031", "Barbara Allen",     date(2026, 6, 10), "Missed: Osteoporosis Check",  None,              None,        "No show. Age 76. On bisphosphonate therapy."),
    ("C032", "P032", "Mark Scott",        date(2026, 6, 24), "Missed: Graves Follow-up",    None,              None,        "No show. TSH was undetectable at last visit."),
    ("C033", "P033", "Nancy Green",       date(2026, 6, 19), "Missed: Adrenal Follow-up",   None,              None,        "No show. On hydrocortisone replacement therapy."),
    # MEDICATION TITRATIONS (8)
    ("C034", "P034", "Paul Baker",        date(2026, 6, 16), "Medication Review",           None,              None,        "Insulin glargine increased from 28 to 32 units nightly."),
    ("C035", "P035", "Donna Nelson",      date(2026, 6, 23), "Medication Review",           None,              None,        "Levothyroxine adjusted from 88mcg to 100mcg."),
    ("C036", "P036", "Steven Carter",     date(2026, 6, 18), "Medication Review",           None,              None,        "Metformin increased from 1000mg to 1500mg daily."),
    ("C037", "P037", "Karen Mitchell",    date(2026, 6, 25), "Medication Review",           None,              None,        "Ozempic titrated from 0.5mg to 1.0mg weekly."),
    ("C038", "P038", "Richard Perez",     date(2026, 6, 11), "Medication Review",           None,              None,        "Insulin lispro ICR adjusted from 1:10 to 1:8."),
    ("C039", "P039", "Betty Roberts",     date(2026, 6, 27), "Medication Review",           None,              None,        "Hydrocortisone dose increased 20mg to 25mg/day. Age 73."),
    ("C040", "P040", "Charles Turner",    date(2026, 6, 14), "Medication Review",           None,              None,        "Levothyroxine uptitrated from 50mcg to 75mcg."),
    ("C041", "P041", "Susan Phillips",    date(2026, 6, 29), "Medication Review",           None,              None,        "Jardiance 10mg added to existing Metformin regimen."),
    # OVERDUE MONITORING (5)
    ("C042", "P042", "Donald Campbell",   date(2025, 11, 4), "Overdue Check",               "HbA1c",           None,        "Last HbA1c Nov 2025. Due every 3 months. Now 8 months overdue."),
    ("C043", "P043", "Lisa Parker",       date(2025, 1, 15), "Overdue Check",               "TSH",             None,        "TSH not checked since Jan 2025. On Levothyroxine. 18 months overdue."),
    ("C044", "P044", "Steven Evans",      date(2025, 6, 20), "Overdue Check",               None,              None,        "Annual diabetic eye and foot exam not completed. 12 months overdue."),
    ("C045", "P045", "Mary Edwards",      date(2023, 3, 10), "Overdue Check",               None,              None,        "DEXA scan last done Mar 2023. Age 71. 3 years overdue."),
    ("C046", "P046", "Anthony Collins",   date(2025, 5, 8),  "Overdue Check",               "HbA1c",           None,        "HbA1c and lipid panel overdue. Last labs May 2025."),
    # NEW DIAGNOSES (4)
    ("C047", "P047", "Helen Stewart",     date(2026, 6, 30), "New Diagnosis Consult",       "Fasting Glucose", "198 mg/dL", "Type 2 Diabetes newly diagnosed. Diabetes education referral needed."),
    ("C048", "P048", "Frank Sanchez",     date(2026, 6, 28), "New Diagnosis Consult",       "TSH",             "9.8 mIU/L", "Hypothyroidism newly diagnosed. Levothyroxine 50mcg started."),
    ("C049", "P049", "Ruth Morris",       date(2026, 6, 26), "New Diagnosis Consult",       None,              None,        "PCOS newly diagnosed. Age 32. Lifestyle counseling given."),
    ("C050", "P050", "Raymond Rogers",    date(2026, 6, 24), "New Diagnosis Consult",       "Calcium",         "11.8 mg/dL","Primary hyperparathyroidism newly diagnosed. Age 67. Surgery referral pending."),
]
# fmt: on


def create_seed_workbook(path: str | Path) -> None:
    """Write a seed master.xlsx at the given path."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default Sheet

    _build_candidates_tab(wb)
    _build_empty_tab(wb, TAB_DRAFTS, DRAFTS_COLUMNS)
    _build_empty_tab(wb, TAB_MANUAL_QUEUE, MANUAL_QUEUE_COLUMNS)
    _build_empty_tab(wb, TAB_LOG, LOG_COLUMNS)
    _build_config_tab(wb)

    wb.save(path)


def _header_row(ws, columns: list[str]) -> None:
    ws.append(columns)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold


def _build_candidates_tab(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet(TAB_CANDIDATES)
    _header_row(ws, CANDIDATES_ALL_COLUMNS)

    for row_data in _CANDIDATE_ROWS:
        candidate_id, patient_id, patient_name, appt_date, visit_type, lab_type, lab_value, notes = row_data
        ws.append([
            candidate_id,
            patient_id,
            patient_name,
            appt_date.isoformat() if appt_date else None,
            visit_type,
            lab_type,
            lab_value,
            notes,
            # Agent-filled columns — left blank to simulate unprocessed admin entry
            None,  # trigger_reason
            None,  # flags
            None,  # urgency_tier
            None,  # channel
            None,  # status  (reader coerces None → "New")
            None,  # last_updated
            "endo",  # specialty_id — set by admin, overridable by agent
        ])


def _build_empty_tab(wb: openpyxl.Workbook, title: str, columns: list[str]) -> None:
    ws = wb.create_sheet(title)
    _header_row(ws, columns)


def _build_config_tab(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet(TAB_CONFIG)
    _header_row(ws, CONFIG_COLUMNS)

    runtime_settings = [
        ("autosend_enabled", "False"),
        ("specialty_scope", "endo"),
        ("classifier_model", "claude-sonnet-4-6"),
        ("triage_model", "claude-sonnet-4-6"),
        ("temperature", "0.0"),
    ]
    for key, value in runtime_settings:
        ws.append([key, value])

    ws.append([None, None])  # blank separator row

    for flag_name, definition in FLAG_VOCABULARY.items():
        ws.append([f"flag:{flag_name}", definition])


if __name__ == "__main__":
    output = sys.argv[1] if len(sys.argv) > 1 else "master.xlsx"
    create_seed_workbook(output)
    print(f"Created {output} with {len(_CANDIDATE_ROWS)} seed candidates.")
